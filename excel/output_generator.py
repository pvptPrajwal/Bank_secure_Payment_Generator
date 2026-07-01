"""
excel/output_generator.py
--------------------------
Generates the two-sheet protected output workbook:
  VALID_PAYMENTS   — matched rows with bank details
  INVALID_PAYMENTS — unmatched rows with reason
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.protection import SheetProtection

VALID_COLS   = ["Party-Name", "Party-Code", "Bank-Account-No", "IFSC-Code", "Amount"]
INVALID_COLS = ["Party-Name", "Party-Code", "Amount", "Reason"]

GREEN  = PatternFill(start_color="1F6F43", end_color="1F6F43", fill_type="solid")
RED    = PatternFill(start_color="8C1D1D", end_color="8C1D1D", fill_type="solid")
WHITE  = Font(color="FFFFFF", bold=True, size=11)
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"),
)
SHEET_PASSWORD = "Bank@2026"


def _write_sheet(ws, columns, rows, fill):
    ws.append(columns)
    for i, _ in enumerate(columns, 1):
        c = ws.cell(1, i)
        c.font, c.fill = WHITE, fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    for row in rows:
        ws.append([row.get(col, "") for col in columns])

    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(columns)):
        for c in r:
            c.border = BORDER
            c.alignment = Alignment(vertical="center")

    for i, col in enumerate(columns, 1):
        max_len = max((len(str(ws.cell(r, i).value or "")) for r in range(1, ws.max_row + 1)), default=len(col))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    ws.protection = SheetProtection(
        sheet=True, password=SHEET_PASSWORD,
        formatCells=False, insertColumns=False, insertRows=False,
        deleteColumns=False, deleteRows=False, sort=False,
        autoFilter=False, selectLockedCells=False, selectUnlockedCells=False,
    )


def generate_output_file(valid_rows: list, invalid_rows: list, output_path: str) -> str:
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb = Workbook()

    ws_v = wb.active
    ws_v.title = "VALID_PAYMENTS"
    _write_sheet(ws_v, VALID_COLS, valid_rows, GREEN)

    ws_i = wb.create_sheet("INVALID_PAYMENTS")
    _write_sheet(ws_i, INVALID_COLS, invalid_rows, RED)

    wb.security = WorkbookProtection(workbookPassword=SHEET_PASSWORD, lockStructure=True)
    wb.save(output_path)
    return output_path


def suggested_output_path() -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(os.path.expanduser("~"), "Desktop", f"BankPaymentOutput_{ts}.xlsx")
